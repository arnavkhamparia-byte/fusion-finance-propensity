"""
Generate Excel report from propensity_results.json.

3 sheets — High / Medium / Low — each sorted by rank.
Output: data/propensity_report.xlsx
"""

import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE  = os.path.join(BASE_DIR, "data", "propensity_results.json")
OUTPUT_FILE = os.path.join(BASE_DIR, "data", "propensity_report.xlsx")

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("Rank",                  "rank"),
    ("Loan Number",           "loan_number"),
    ("Customer Name",         "name"),
    ("City",                  "city"),
    ("Propensity Score",      "propensity_score"),
    ("DPD Bucket",            "dpd_bucket"),
    ("Loan Amount (₹)",       "loan_amount"),
    ("Principal Outstanding (₹)", "principal_outstanding"),
    ("EMI Amount (₹)",        "emi_amount"),
    ("Primary Contact",       "primary_contact_number"),
    ("WhatsApp",              "whatsapp_contact_number"),
    ("Co-Applicant Name",     "co_applicant_name"),
    ("Co-Applicant Contact",  "co_applicant_contact"),
    ("Reference Contact 1",   "reference_contact_1"),
    ("Reference Contact 2",   "reference_contact_2"),
    ("Disposition",           "disposition"),
    ("Commitment Strength",   "commitment_strength"),
    ("Engagement Level",      "engagement_level"),
    ("Promise Made",          "promise_made"),
    ("Promise Date",          "promise_date"),
    ("Sentiment",             "sentiment"),
    ("Key Reasons",           "key_reasons"),
    ("AI Summary",            "summary"),
    ("Recording Mismatch",    "recording_mismatch"),
    ("Analysed At",           "analysed_at"),
]

# ── Tier theme colours ────────────────────────────────────────────────────────
TIER_THEMES = {
    "High":   {"header_fill": "1a6b3a", "header_font": "ffffff", "tab": "22c55e"},
    "Medium": {"header_fill": "92400e", "header_font": "ffffff", "tab": "f97316"},
    "Low":    {"header_fill": "7f1d1d", "header_font": "ffffff", "tab": "ef4444"},
}

# Score gradient: green (high) → orange → red (low)
def score_fill(score):
    if score >= 65:
        return PatternFill("solid", fgColor="d1fae5")   # light green
    elif score >= 40:
        return PatternFill("solid", fgColor="ffedd5")   # light orange
    else:
        return PatternFill("solid", fgColor="fee2e2")   # light red

def thin_border():
    s = Side(style="thin", color="d1d5db")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_value(col_key, value):
    """Format cell value for display."""
    if value is None:
        return "—"
    if col_key == "promise_made" or col_key == "recording_mismatch":
        return "Yes" if value else "No"
    if col_key == "key_reasons" and isinstance(value, list):
        return "; ".join(value)
    if col_key in ("loan_amount", "principal_outstanding", "emi_amount"):
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    if col_key == "analysed_at":
        try:
            dt = datetime.fromisoformat(str(value))
            return dt.strftime("%d %b %Y, %I:%M %p")
        except Exception:
            return str(value)
    return value


def write_sheet(wb, tier, accounts):
    ws = wb.create_sheet(title=f"{tier} Propensity")
    theme = TIER_THEMES[tier]

    # Tab colour
    ws.sheet_properties.tabColor = theme["tab"]

    # ── Header row ────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=theme["header_fill"])
    header_font = Font(bold=True, color=theme["header_font"], size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_label)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border()

    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, account in enumerate(accounts, 2):
        score = account.get("propensity_score", 0)
        row_fill = score_fill(score)

        for col_idx, (col_label, col_key) in enumerate(COLUMNS, 1):
            raw   = account.get(col_key)
            value = fmt_value(col_key, raw)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = thin_border()
            cell.font   = Font(size=9)

            # Alignment
            if col_key in ("rank", "propensity_score"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_key in ("loan_amount", "principal_outstanding", "emi_amount"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            elif col_key in ("key_reasons", "summary"):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Row height — taller for rows with long text
        ws.row_dimensions[row_idx].height = 40

    # ── Column widths ─────────────────────────────────────────────
    col_widths = {
        "Rank": 6,
        "Loan Number": 16,
        "Customer Name": 22,
        "City": 14,
        "Propensity Score": 12,
        "DPD Bucket": 20,
        "Loan Amount (₹)": 16,
        "Principal Outstanding (₹)": 20,
        "EMI Amount (₹)": 14,
        "Primary Contact": 16,
        "WhatsApp": 16,
        "Co-Applicant Name": 20,
        "Co-Applicant Contact": 18,
        "Reference Contact 1": 18,
        "Reference Contact 2": 18,
        "Disposition": 28,
        "Commitment Strength": 16,
        "Engagement Level": 14,
        "Promise Made": 12,
        "Promise Date": 14,
        "Sentiment": 12,
        "Key Reasons": 45,
        "AI Summary": 50,
        "Recording Mismatch": 16,
        "Analysed At": 20,
    }

    for col_idx, (col_label, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_label, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Summary row at bottom ─────────────────────────────────────
    summary_row = len(accounts) + 3
    ws.cell(row=summary_row, column=1, value=f"Total: {len(accounts)} accounts").font = Font(bold=True, size=10)

    print(f"  Sheet '{tier} Propensity' — {len(accounts)} accounts written")


def main():
    with open(INPUT_FILE) as f:
        data = json.load(f)

    accounts = data.get("accounts", [])

    # Split by tier, preserving rank order
    tiers = {
        "High":   [a for a in accounts if a.get("tier") == "High"],
        "Medium": [a for a in accounts if a.get("tier") == "Medium"],
        "Low":    [a for a in accounts if a.get("tier") == "Low"],
    }

    print(f"\nGenerating Excel report for {len(accounts)} accounts...")
    print(f"  High: {len(tiers['High'])}  |  Medium: {len(tiers['Medium'])}  |  Low: {len(tiers['Low'])}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for tier in ("High", "Medium", "Low"):
        write_sheet(wb, tier, tiers[tier])

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"\nSaved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
